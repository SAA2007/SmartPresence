import sqlite3
import os
import json
import pickle
import base64
import requests as http_requests
import numpy as np
from flask import Blueprint, jsonify, request, send_file, current_app, session
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from io import BytesIO
from functools import wraps

api_bp = Blueprint('api', __name__, url_prefix='/api')


def get_db():
    db_path = current_app.config['DB_PATH']
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


# ── Auth Helpers ──

def api_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({"error": "Authentication required"}), 401
        return f(*args, **kwargs)
    return decorated


def api_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({"error": "Authentication required"}), 401
        if session.get('role') != 'admin':
            return jsonify({"error": "Admin access required"}), 403
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════════════════
#  AUTH — Login / Logout / Session
# ══════════════════════════════════════════════════════

@api_bp.route('/auth/login', methods=['POST'])
def auth_login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400

    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()

    if not user or not check_password_hash(user['password_hash'], password):
        return jsonify({"error": "Invalid username or password"}), 401

    session['user_id'] = user['id']
    session['username'] = user['username']
    session['display_name'] = user['display_name']
    session['role'] = user['role']
    session.permanent = True

    return jsonify({
        "success": True,
        "user": {
            "id": user['id'],
            "username": user['username'],
            "display_name": user['display_name'],
            "role": user['role']
        }
    })


@api_bp.route('/auth/logout', methods=['POST'])
def auth_logout():
    session.clear()
    return jsonify({"success": True})


@api_bp.route('/auth/me', methods=['GET'])
def auth_me():
    if 'user_id' not in session:
        return jsonify({"authenticated": False}), 401
    return jsonify({
        "authenticated": True,
        "user": {
            "id": session['user_id'],
            "username": session['username'],
            "display_name": session['display_name'],
            "role": session['role']
        }
    })


@api_bp.route('/auth/verify-pin', methods=['POST'])
def verify_settings_pin():
    data = request.get_json()
    pin = data.get('pin', '')
    from ai_module import common
    if pin == common.SETTINGS_PIN:
        session['settings_unlocked'] = True
        return jsonify({"success": True})
    return jsonify({"error": "Invalid PIN"}), 401


# ══════════════════════════════════════════════════════
#  USERS — Admin CRUD
# ══════════════════════════════════════════════════════

@api_bp.route('/users', methods=['GET'])
@api_admin_required
def get_users():
    conn = get_db()
    users = conn.execute("SELECT id, username, display_name, role, created_at FROM users ORDER BY created_at").fetchall()
    conn.close()
    return jsonify([dict(u) for u in users])


@api_bp.route('/users', methods=['POST'])
@api_admin_required
def add_user():
    data = request.get_json()
    username = data.get('username', '').strip()
    display_name = data.get('display_name', '').strip()
    password = data.get('password', '')
    role = data.get('role', 'teacher')

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400
    if len(password) < 4:
        return jsonify({"error": "Password must be at least 4 characters"}), 400
    if role not in ('admin', 'teacher'):
        return jsonify({"error": "Role must be 'admin' or 'teacher'"}), 400

    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO users (username, display_name, password_hash, role) VALUES (?, ?, ?, ?)",
            (username, display_name or username, generate_password_hash(password), role))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"error": "Username already exists"}), 409
    conn.close()
    return jsonify({"success": True}), 201


@api_bp.route('/users/<int:uid>', methods=['PUT'])
@api_admin_required
def update_user(uid):
    data = request.get_json()
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (uid,)).fetchone()
    if not user:
        conn.close()
        return jsonify({"error": "User not found"}), 404

    display_name = data.get('display_name', user['display_name']).strip()
    role = data.get('role', user['role'])

    if role not in ('admin', 'teacher'):
        conn.close()
        return jsonify({"error": "Invalid role"}), 400

    conn.execute("UPDATE users SET display_name=?, role=? WHERE id=?", (display_name, role, uid))

    # Update password if provided
    new_password = data.get('password', '').strip()
    if new_password:
        if len(new_password) < 4:
            conn.close()
            return jsonify({"error": "Password must be at least 4 characters"}), 400
        conn.execute("UPDATE users SET password_hash=? WHERE id=?",
                     (generate_password_hash(new_password), uid))

    conn.commit()
    conn.close()
    return jsonify({"success": True})


@api_bp.route('/users/<int:uid>', methods=['DELETE'])
@api_admin_required
def delete_user(uid):
    if uid == session.get('user_id'):
        return jsonify({"error": "Cannot delete your own account"}), 400
    conn = get_db()
    conn.execute("DELETE FROM users WHERE id = ?", (uid,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ══════════════════════════════════════════════════════
#  STUDENTS — CRUD
# ══════════════════════════════════════════════════════

@api_bp.route('/students', methods=['GET'])
@api_login_required
def get_students():
    conn = get_db()
    students = conn.execute("SELECT * FROM students ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(s) for s in students])


@api_bp.route('/students', methods=['POST'])
@api_login_required
def add_student():
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({"error": "Name is required"}), 400
    conn = get_db()
    try:
        conn.execute("INSERT INTO students (name, student_id, email, notes) VALUES (?, ?, ?, ?)",
                     (name, data.get('student_id', '').strip(), data.get('email', '').strip(),
                      data.get('notes', '').strip()))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"error": "Student already exists"}), 409
    conn.close()
    return jsonify({"success": True, "name": name}), 201


@api_bp.route('/students/<int:sid>', methods=['GET'])
@api_login_required
def get_student(sid):
    conn = get_db()
    student = conn.execute("SELECT * FROM students WHERE id = ?", (sid,)).fetchone()
    if not student:
        conn.close()
        return jsonify({"error": "Student not found"}), 404

    # Include attendance summary
    total_classes = conn.execute(
        "SELECT COUNT(*) FROM attendance_logs WHERE student_id = ?", (sid,)).fetchone()[0]
    present_count = conn.execute(
        "SELECT COUNT(*) FROM attendance_logs WHERE student_id = ? AND status IN ('Present', 'On Time')",
        (sid,)).fetchone()[0]
    late_count = conn.execute(
        "SELECT COUNT(*) FROM attendance_logs WHERE student_id = ? AND status = 'Late'",
        (sid,)).fetchone()[0]
    absent_count = conn.execute(
        "SELECT COUNT(*) FROM attendance_logs WHERE student_id = ? AND status = 'Absent'",
        (sid,)).fetchone()[0]

    # Recent attendance
    recent = conn.execute("""
        SELECT al.id, al.timestamp, al.status, al.source, al.notes
        FROM attendance_logs al WHERE al.student_id = ?
        ORDER BY al.timestamp DESC LIMIT 50
    """, (sid,)).fetchall()

    conn.close()

    attendance_rate = round((present_count + late_count) / total_classes * 100, 1) if total_classes > 0 else 0

    result = dict(student)
    result['stats'] = {
        'total_classes': total_classes,
        'present': present_count,
        'late': late_count,
        'absent': absent_count,
        'attendance_rate': attendance_rate
    }
    result['recent_attendance'] = [dict(r) for r in recent]
    return jsonify(result)


@api_bp.route('/students/<int:sid>', methods=['PUT'])
@api_login_required
def update_student(sid):
    data = request.get_json()
    conn = get_db()
    student = conn.execute("SELECT * FROM students WHERE id = ?", (sid,)).fetchone()
    if not student:
        conn.close()
        return jsonify({"error": "Student not found"}), 404
    conn.execute("UPDATE students SET name=?, student_id=?, email=?, notes=? WHERE id=?",
                 (data.get('name', student['name']).strip(),
                  data.get('student_id', student['student_id'] or '').strip(),
                  data.get('email', student['email'] or '').strip(),
                  data.get('notes', student['notes'] or '').strip(), sid))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@api_bp.route('/students/<int:sid>', methods=['DELETE'])
@api_login_required
def delete_student(sid):
    conn = get_db()
    conn.execute("DELETE FROM attendance_logs WHERE student_id = ?", (sid,))
    conn.execute("DELETE FROM students WHERE id = ?", (sid,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ══════════════════════════════════════════════════════
#  ENROLLMENT — Web-based
# ══════════════════════════════════════════════════════

@api_bp.route('/enroll', methods=['POST'])
@api_login_required
def enroll_student():
    data = request.get_json()
    name = data.get('name', '').strip()
    image_b64 = data.get('image', '')
    if not name:
        return jsonify({"error": "Name is required"}), 400
    if not image_b64:
        return jsonify({"error": "Image is required"}), 400

    try:
        import face_recognition
        import cv2

        if ',' in image_b64:
            image_b64 = image_b64.split(',')[1]
        image_bytes = base64.b64decode(image_b64)
        np_arr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
        if img is None:
            return jsonify({"error": "Invalid image data"}), 400

        rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        boxes = face_recognition.face_locations(rgb, model="hog")
        if len(boxes) == 0:
            return jsonify({"error": "No face detected. Try a clearer photo."}), 400
        if len(boxes) > 1:
            return jsonify({"error": "Multiple faces detected. Use a photo with only one person."}), 400

        encodings = face_recognition.face_encodings(rgb, boxes)
        if not encodings:
            return jsonify({"error": "Could not encode the face."}), 500
        new_encoding = encodings[0]

        from ai_module import common
        enc_path = common.ENCODINGS_PATH
        if os.path.exists(enc_path):
            with open(enc_path, "rb") as f:
                enc_data = pickle.load(f)
        else:
            enc_data = {"names": [], "encodings": []}

        enc_data["names"].append(name)
        enc_data["encodings"].append(new_encoding)
        with open(enc_path, "wb") as f:
            pickle.dump(enc_data, f)

        conn = get_db()
        try:
            conn.execute("INSERT INTO students (name, student_id, email) VALUES (?, ?, ?)",
                         (name, data.get('student_id', '').strip(), data.get('email', '').strip()))
            conn.commit()
        except sqlite3.IntegrityError:
            pass
        conn.close()

        try:
            from web_app.video_stream import video_stream
            if video_stream.is_running:
                video_stream.face_system.known_names.append(name)
                video_stream.face_system.known_encodings.append(new_encoding)
        except Exception:
            pass

        return jsonify({"success": True, "name": name}), 201
    except ImportError as e:
        return jsonify({"error": f"AI library not available: {e}"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════
#  ATTENDANCE — Read + Manual Entry/Override
# ══════════════════════════════════════════════════════

@api_bp.route('/attendance', methods=['GET'])
@api_login_required
def get_attendance():
    date_filter = request.args.get('date')
    student_id = request.args.get('student_id')
    conn = get_db()

    if student_id:
        rows = conn.execute("""
            SELECT al.id, s.name, al.timestamp, al.status, al.source, al.notes
            FROM attendance_logs al JOIN students s ON al.student_id = s.id
            WHERE al.student_id = ? ORDER BY al.timestamp DESC LIMIT 200
        """, (student_id,)).fetchall()
    elif date_filter:
        rows = conn.execute("""
            SELECT al.id, s.name, al.timestamp, al.status, al.source, al.notes
            FROM attendance_logs al JOIN students s ON al.student_id = s.id
            WHERE DATE(al.timestamp) = ? ORDER BY al.timestamp DESC
        """, (date_filter,)).fetchall()
    else:
        rows = conn.execute("""
            SELECT al.id, s.name, al.timestamp, al.status, al.source, al.notes
            FROM attendance_logs al JOIN students s ON al.student_id = s.id
            ORDER BY al.timestamp DESC LIMIT 200
        """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@api_bp.route('/attendance', methods=['POST'])
@api_login_required
def manual_attendance():
    data = request.get_json()
    student_id = data.get('student_id')
    status = data.get('status', 'Present')
    notes = data.get('notes', '').strip()
    if not student_id:
        return jsonify({"error": "student_id is required"}), 400

    from ai_module import common
    if status not in common.VALID_STATUSES:
        return jsonify({"error": f"Invalid status. Must be one of: {common.VALID_STATUSES}"}), 400

    conn = get_db()
    student = conn.execute("SELECT id FROM students WHERE id = ?", (student_id,)).fetchone()
    if not student:
        conn.close()
        return jsonify({"error": "Student not found"}), 404

    conn.execute("INSERT INTO attendance_logs (student_id, status, source, notes) VALUES (?, ?, 'manual', ?)",
                 (student_id, status, notes))
    conn.commit()
    conn.close()
    return jsonify({"success": True}), 201


@api_bp.route('/attendance/<int:log_id>', methods=['PUT'])
@api_login_required
def override_attendance(log_id):
    data = request.get_json()
    conn = get_db()
    log = conn.execute("SELECT * FROM attendance_logs WHERE id = ?", (log_id,)).fetchone()
    if not log:
        conn.close()
        return jsonify({"error": "Attendance record not found"}), 404

    from ai_module import common
    new_status = data.get('status', log['status'])
    if new_status not in common.VALID_STATUSES:
        conn.close()
        return jsonify({"error": f"Invalid status. Must be one of: {common.VALID_STATUSES}"}), 400

    conn.execute("UPDATE attendance_logs SET status=?, notes=?, source='override' WHERE id=?",
                 (new_status, data.get('notes', log['notes'] or ''), log_id))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@api_bp.route('/attendance/<int:log_id>', methods=['DELETE'])
@api_login_required
def delete_attendance(log_id):
    conn = get_db()
    conn.execute("DELETE FROM attendance_logs WHERE id = ?", (log_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ══════════════════════════════════════════════════════
#  SCHEDULE — Timetable CRUD
# ══════════════════════════════════════════════════════

@api_bp.route('/schedule', methods=['GET'])
@api_login_required
def get_schedules():
    conn = get_db()
    rows = conn.execute("SELECT * FROM class_schedules ORDER BY day_of_week, start_time").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@api_bp.route('/schedule', methods=['POST'])
@api_login_required
def add_schedule():
    data = request.get_json()
    day = data.get('day_of_week', '').strip()
    start = data.get('start_time', '').strip()
    end = data.get('end_time', '').strip()
    name = data.get('class_name', 'Class').strip()

    if not all([day, start, end]):
        return jsonify({"error": "day_of_week, start_time, end_time are required"}), 400

    valid_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    if day not in valid_days:
        return jsonify({"error": f"Invalid day. Must be one of: {valid_days}"}), 400

    # Validate end > start
    if end <= start:
        return jsonify({"error": "End time must be after start time"}), 400

    conn = get_db()
    conn.execute("INSERT INTO class_schedules (day_of_week, start_time, end_time, class_name) VALUES (?, ?, ?, ?)",
                 (day, start, end, name))
    conn.commit()
    conn.close()
    return jsonify({"success": True}), 201


@api_bp.route('/schedule/<int:sid>', methods=['PUT'])
@api_login_required
def update_schedule(sid):
    data = request.get_json()
    conn = get_db()
    sched = conn.execute("SELECT * FROM class_schedules WHERE id = ?", (sid,)).fetchone()
    if not sched:
        conn.close()
        return jsonify({"error": "Schedule not found"}), 404

    conn.execute("UPDATE class_schedules SET day_of_week=?, start_time=?, end_time=?, class_name=?, is_active=? WHERE id=?",
                 (data.get('day_of_week', sched['day_of_week']),
                  data.get('start_time', sched['start_time']),
                  data.get('end_time', sched['end_time']),
                  data.get('class_name', sched['class_name']),
                  data.get('is_active', sched['is_active']),
                  sid))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@api_bp.route('/schedule/<int:sid>', methods=['DELETE'])
@api_login_required
def delete_schedule(sid):
    conn = get_db()
    conn.execute("DELETE FROM class_schedules WHERE id = ?", (sid,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ══════════════════════════════════════════════════════
#  SYSTEM — Start / Stop / Restart / Mode / Status
# ══════════════════════════════════════════════════════

@api_bp.route('/system/status', methods=['GET'])
def system_status():
    from web_app.video_stream import video_stream
    return jsonify(video_stream.get_status())


@api_bp.route('/system', methods=['POST'])
@api_login_required
def system_control():
    data = request.get_json()
    action = data.get('action', '')
    from web_app.video_stream import video_stream
    from ai_module import common

    if action == 'start':
        video_stream.start()
        return jsonify({"success": True, "message": "AI started"})
    elif action == 'stop':
        video_stream.stop()
        return jsonify({"success": True, "message": "AI stopped"})
    elif action == 'restart':
        video_stream.restart()
        return jsonify({"success": True, "message": "AI restarted"})
    elif action == 'shutdown':
        if session.get('role') != 'admin':
            return jsonify({"error": "Only admins can shutdown the server"}), 403
        video_stream.stop()
        import threading
        def _shutdown():
            import time
            time.sleep(1)
            os._exit(0)
        threading.Thread(target=_shutdown, daemon=True).start()
        return jsonify({"success": True, "message": "Server shutting down..."})
    elif action == 'set_mode':
        mode = data.get('mode', 'auto')
        if mode not in ('auto', 'force_on', 'force_off'):
            return jsonify({"error": "Invalid mode"}), 400
        common.SYSTEM_MODE = mode
        return jsonify({"success": True, "mode": mode})
    else:
        return jsonify({"error": "Invalid action. Use: start, stop, restart, shutdown, set_mode"}), 400


# ══════════════════════════════════════════════════════
#  CRASH REPORTING + TELEGRAM
# ══════════════════════════════════════════════════════

def send_telegram(message):
    """Send a message to the configured Telegram chat."""
    from ai_module import common
    token = common.TELEGRAM_BOT_TOKEN
    chat_id = common.TELEGRAM_CHAT_ID

    if not token or token == 'YOUR_BOT_TOKEN_HERE' or not chat_id or chat_id == 'YOUR_CHAT_ID_HERE':
        return False

    try:
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        http_requests.post(url, json={
            "chat_id": chat_id,
            "text": message,
            "parse_mode": "Markdown"
        }, timeout=5)
        return True
    except Exception:
        return False


@api_bp.route('/report', methods=['POST'])
@api_login_required
def submit_report():
    data = request.get_json()
    description = data.get('description', '').strip()
    if not description:
        return jsonify({"error": "Description is required"}), 400

    from ai_module import common
    from web_app.video_stream import video_stream

    category = data.get('category', 'bug')
    severity = data.get('severity', 'medium')

    # Build report
    report = {
        "timestamp": datetime.now().isoformat(),
        "description": description,
        "category": category,
        "severity": severity,
        "submitted_by": session.get('display_name', 'Unknown'),
    }

    # Selectable data inclusions
    if data.get('include_system', True):
        report["system_state"] = video_stream.get_status()
    if data.get('include_settings', True):
        report["settings"] = {
            "detection_scale": getattr(common, 'DETECTION_SCALE', 0.5),
            "tolerance": getattr(common, 'TOLERANCE', 0.5),
            "late_threshold": getattr(common, 'LATE_THRESHOLD', 10),
            "disappear_threshold": getattr(common, 'DISAPPEAR_THRESHOLD', 15),
            "system_mode": getattr(common, 'SYSTEM_MODE', 'auto'),
        }
    if data.get('include_browser', False):
        report["browser"] = data.get('user_agent', '')

    # Save locally
    reports_dir = common.CRASH_REPORTS_DIR
    os.makedirs(reports_dir, exist_ok=True)
    filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    filepath = os.path.join(reports_dir, filename)
    with open(filepath, 'w') as f:
        json.dump(report, f, indent=2, default=str)

    # Send to Telegram
    severity_emoji = {'low': '🟡', 'medium': '🟠', 'critical': '🔴'}.get(severity, '⚪')
    tg_msg = (
        f"{severity_emoji} *SmartPresence Report*\n"
        f"*Category:* {category}\n"
        f"*Severity:* {severity}\n"
        f"*By:* {session.get('display_name', 'Unknown')}\n\n"
        f"_{description}_\n\n"
        f"📁 Saved: `{filename}`"
    )
    telegram_sent = send_telegram(tg_msg)

    return jsonify({
        "success": True,
        "file": filename,
        "telegram_sent": telegram_sent
    }), 201


# ══════════════════════════════════════════════════════
#  STATS + CHART DATA
# ══════════════════════════════════════════════════════

@api_bp.route('/stats', methods=['GET'])
@api_login_required
def get_stats():
    conn = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    total_students = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    present_today = conn.execute("""
        SELECT COUNT(DISTINCT student_id) FROM attendance_logs
        WHERE DATE(timestamp) = ? AND status IN ('Present', 'On Time', 'Late')
    """, (today,)).fetchone()[0]
    absent_today = total_students - present_today
    total_logs = conn.execute("SELECT COUNT(*) FROM attendance_logs").fetchone()[0]
    conn.close()
    return jsonify({
        "total_students": total_students,
        "present_today": present_today,
        "absent_today": absent_today,
        "total_logs": total_logs,
        "date": today
    })


@api_bp.route('/stats/chart', methods=['GET'])
@api_login_required
def chart_data():
    days = int(request.args.get('days', 7))
    conn = get_db()
    labels, present_data, absent_data = [], [], []
    total_students = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]

    for i in range(days - 1, -1, -1):
        date = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        labels.append(date)
        present = conn.execute("""
            SELECT COUNT(DISTINCT student_id) FROM attendance_logs
            WHERE DATE(timestamp) = ? AND status IN ('Present', 'On Time', 'Late')
        """, (date,)).fetchone()[0]
        present_data.append(present)
        absent_data.append(max(0, total_students - present))

    conn.close()
    return jsonify({"labels": labels, "present": present_data,
                    "absent": absent_data, "total_students": total_students})


# ══════════════════════════════════════════════════════
#  EXPORT
# ══════════════════════════════════════════════════════

@api_bp.route('/export', methods=['GET'])
@api_login_required
def export_data():
    fmt = request.args.get('format', 'xlsx')
    date_filter = request.args.get('date')
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    conn = get_db()

    if date_from and date_to:
        rows = conn.execute("""
            SELECT s.name, s.student_id as sid, al.timestamp, al.status, al.source, al.notes
            FROM attendance_logs al JOIN students s ON al.student_id = s.id
            WHERE DATE(al.timestamp) BETWEEN ? AND ? ORDER BY al.timestamp
        """, (date_from, date_to)).fetchall()
        date_label = f"{date_from}_to_{date_to}"
    elif date_filter:
        rows = conn.execute("""
            SELECT s.name, s.student_id as sid, al.timestamp, al.status, al.source, al.notes
            FROM attendance_logs al JOIN students s ON al.student_id = s.id
            WHERE DATE(al.timestamp) = ? ORDER BY al.timestamp
        """, (date_filter,)).fetchall()
        date_label = date_filter
    else:
        rows = conn.execute("""
            SELECT s.name, s.student_id as sid, al.timestamp, al.status, al.source, al.notes
            FROM attendance_logs al JOIN students s ON al.student_id = s.id
            ORDER BY al.timestamp DESC LIMIT 1000
        """).fetchall()
        date_label = "all"
    conn.close()

    return _export_csv(rows, date_label) if fmt == 'csv' else _export_xlsx(rows, date_label)


def _export_xlsx(rows, date_label):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
    headers = ["Student Name", "Student ID", "Timestamp", "Status", "Source", "Notes"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row['name'], row['sid'] or '', row['timestamp'],
                   row['status'], row['source'], row['notes'] or ''])
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"attendance_{date_label}.xlsx",
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _export_csv(rows, date_label):
    import csv
    from io import StringIO
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Student Name", "Student ID", "Timestamp", "Status", "Source", "Notes"])
    for row in rows:
        writer.writerow([row['name'], row['sid'] or '', row['timestamp'],
                         row['status'], row['source'], row['notes'] or ''])
    csv_bytes = BytesIO(output.getvalue().encode('utf-8'))
    csv_bytes.seek(0)
    return send_file(csv_bytes, download_name=f"attendance_{date_label}.csv",
                     as_attachment=True, mimetype='text/csv')


# ══════════════════════════════════════════════════════
#  SETTINGS
# ══════════════════════════════════════════════════════

@api_bp.route('/settings', methods=['GET'])
@api_login_required
def get_settings():
    try:
        from ai_module import common
        return jsonify({
            "detection_scale": getattr(common, 'DETECTION_SCALE', 0.5),
            "tolerance": getattr(common, 'TOLERANCE', 0.5),
            "frame_width": getattr(common, 'FRAME_WIDTH', 1920),
            "frame_height": getattr(common, 'FRAME_HEIGHT', 1080),
            "late_threshold": getattr(common, 'LATE_THRESHOLD', 10),
            "disappear_threshold": getattr(common, 'DISAPPEAR_THRESHOLD', 15),
            "system_mode": getattr(common, 'SYSTEM_MODE', 'auto'),
        })
    except ImportError:
        return jsonify({"error": "AI module not found"}), 500


@api_bp.route('/settings', methods=['POST'])
@api_login_required
def update_settings():
    data = request.get_json()
    try:
        from ai_module import common
        if 'detection_scale' in data:
            common.DETECTION_SCALE = float(data['detection_scale'])
        if 'tolerance' in data:
            common.TOLERANCE = float(data['tolerance'])
        if 'late_threshold' in data:
            common.LATE_THRESHOLD = int(data['late_threshold'])
        if 'disappear_threshold' in data:
            common.DISAPPEAR_THRESHOLD = int(data['disappear_threshold'])
        if 'system_mode' in data:
            common.SYSTEM_MODE = data['system_mode']
        return jsonify({"success": True, "message": "Settings updated"})
    except ImportError:
        return jsonify({"error": "AI module not found"}), 500


# ══════════════════════════════════════════════════════
#  ADMIN CONFIG — .env Editing (PIN-gated)
# ══════════════════════════════════════════════════════

# Editable keys — anything NOT in this list is hidden from the UI
EDITABLE_ENV_KEYS = [
    'SETTINGS_PIN',
    'SECRET_KEY',
    'TELEGRAM_BOT_TOKEN',
    'TELEGRAM_CHAT_ID',
    'DB_ENCRYPTION_KEY',
]

# Keys whose values should be masked in GET responses
MASKED_KEYS = {'SECRET_KEY', 'DB_ENCRYPTION_KEY', 'SETTINGS_PIN'}


def _env_path():
    """Return the absolute path to the project .env file."""
    return os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), '.env')


def _read_env():
    """Read .env file into an ordered list of (key, value, is_comment) tuples."""
    lines = []
    path = _env_path()
    if not os.path.exists(path):
        return lines
    with open(path, 'r', encoding='utf-8') as f:
        for raw in f:
            raw = raw.rstrip('\r\n')
            if raw.startswith('#') or raw.strip() == '':
                lines.append((None, raw, True))
            elif '=' in raw:
                k, v = raw.split('=', 1)
                lines.append((k.strip(), v.strip(), False))
            else:
                lines.append((None, raw, True))
    return lines


def _write_env(lines):
    """Write list of (key, value, is_comment) tuples back to .env."""
    path = _env_path()
    with open(path, 'w', encoding='utf-8', newline='\n') as f:
        for key, val, is_comment in lines:
            if is_comment:
                f.write(val + '\n')
            else:
                f.write(f'{key}={val}\n')


@api_bp.route('/config', methods=['GET'])
@api_admin_required
def get_config():
    """Return editable env variables (masked where appropriate)."""
    if not session.get('settings_unlocked'):
        return jsonify({"error": "Settings PIN required", "pin_required": True}), 403

    lines = _read_env()
    result = {}
    for key, val, is_comment in lines:
        if is_comment or key not in EDITABLE_ENV_KEYS:
            continue
        if key in MASKED_KEYS:
            # Show first 4 + last 4 chars, mask middle
            if len(val) > 10:
                result[key] = val[:4] + '•' * (len(val) - 8) + val[-4:]
            else:
                result[key] = '•' * len(val)
        else:
            result[key] = val

    return jsonify({"config": result, "editable_keys": EDITABLE_ENV_KEYS})


@api_bp.route('/config', methods=['PUT'])
@api_admin_required
def update_config():
    """Update specific .env variables. Only EDITABLE_ENV_KEYS allowed."""
    if not session.get('settings_unlocked'):
        return jsonify({"error": "Settings PIN required", "pin_required": True}), 403

    data = request.get_json()
    updates = data.get('updates', {})

    if not updates:
        return jsonify({"error": "No updates provided"}), 400

    # Validate only editable keys
    for key in updates:
        if key not in EDITABLE_ENV_KEYS:
            return jsonify({"error": f"Key '{key}' is not editable"}), 400

    lines = _read_env()
    existing_keys = {k for k, v, c in lines if not c}

    # Update existing keys
    updated = []
    for key, val, is_comment in lines:
        if not is_comment and key in updates:
            new_val = str(updates[key]).strip()
            if not new_val:
                return jsonify({"error": f"Value for '{key}' cannot be empty"}), 400
            updated.append((key, new_val, False))
        else:
            updated.append((key, val, is_comment))

    # Add new keys that weren't in the file
    for key, val in updates.items():
        if key not in existing_keys:
            new_val = str(val).strip()
            if not new_val:
                return jsonify({"error": f"Value for '{key}' cannot be empty"}), 400
            updated.append((key, new_val, False))

    _write_env(updated)

    # Reload into common module
    try:
        from ai_module import common
        if 'SETTINGS_PIN' in updates:
            common.SETTINGS_PIN = updates['SETTINGS_PIN']
        if 'TELEGRAM_BOT_TOKEN' in updates:
            common.TELEGRAM_BOT_TOKEN = updates['TELEGRAM_BOT_TOKEN']
        if 'TELEGRAM_CHAT_ID' in updates:
            common.TELEGRAM_CHAT_ID = updates['TELEGRAM_CHAT_ID']
    except ImportError:
        pass

    # Update Flask SECRET_KEY if changed
    if 'SECRET_KEY' in updates:
        current_app.config['SECRET_KEY'] = updates['SECRET_KEY']

    return jsonify({"success": True, "message": f"{len(updates)} config value(s) updated", "updated_keys": list(updates.keys())})


@api_bp.route('/config/export-db', methods=['GET'])
@api_admin_required
def export_db():
    """Download a copy of the database file for backup."""
    db_path = current_app.config['DB_PATH']
    if not os.path.exists(db_path):
        return jsonify({"error": "Database not found"}), 404
    return send_file(
        db_path,
        as_attachment=True,
        download_name=f'smartpresence_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db',
        mimetype='application/x-sqlite3'
    )


@api_bp.route('/config/version', methods=['GET'])
@api_login_required
def get_version():
    """Return system version info."""
    return jsonify({
        "version": "1.5.0",
        "phase": "5C",
        "codename": "Security & Polish",
        "build_date": "2026-02-12"
    })

